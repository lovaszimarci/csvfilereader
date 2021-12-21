

import os
import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import workbook 
import csv

wb = load_workbook('TestPlan.xlsx')
ws = wb.active
ls1 = []
for row in range(1,2000):
    for col in range(1, 2):
        char = get_column_letter(col) 
        ls1.append(ws[char + str(row)].value)
wb.close


wb = load_workbook('Full.xlsx')
ws = wb.active
ls2 = []
for row in range(1, 2000):
    for col in range(1,2):
        char = get_column_letter(col)
        ls2.append(ws[char + str(row)].value)
wb.close

setls1 = list(set(ls1))
setls2 = list(set(ls2))


osszlist = setls2 + setls1


for szam in osszlist:
    if osszlist.count(szam) > 1:
        osszlist.remove(szam)
        osszlist.remove(szam)

osszlist = str(osszlist)

with open('3.csv', 'w') as file3:
    file3.write(osszlist)